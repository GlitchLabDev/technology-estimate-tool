import openpyxl, json, re, os, datetime
from openpyxl.utils import get_column_letter

files = [
    ('template', '/home/ubuntu/attachments/6955d04e-8f69-47cc-85a5-edbf4a37e162/Technology_Estimate_Template-71626.xlsx'),
    ('pricing', '/home/ubuntu/attachments/dbb9b6f4-b8bb-427a-89fa-4c88ffebea1f/High_Level_costs_for_Network_kit_1.xlsx'),
]

def convert_value(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()
    return str(v)

def clean_formula(s):
    # remove external workbook link markers
    s = re.sub(r"'\[1\]([^']+)'", r"'\1'", s)
    s = re.sub(r"\[1\]", "", s)
    # patch obvious typo in Estimate Sheet (E341 does not exist, intended E34)
    s = s.replace("'Test Fit'!E341", "'Test Fit'!E34")
    s = re.sub(r'\s+', ' ', s).strip()
    return s

all_sheets = {}
for label, p in files:
    print('Loading', p)
    wb = openpyxl.load_workbook(p, data_only=False)
    for s in wb.sheetnames:
        ws = wb[s]
        cells = {}
        last_row = 0
        last_col = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    last_row = max(last_row, cell.row)
                    last_col = max(last_col, cell.column)
                    cells[(cell.row, cell.column)] = cell
        if last_row == 0:
            continue
        data = []
        for r in range(1, last_row + 1):
            row_arr = []
            for c in range(1, last_col + 1):
                cell = cells.get((r, c))
                if cell is None:
                    row_arr.append(None)
                elif cell.data_type == 'f':
                    row_arr.append(clean_formula(str(cell.value)))
                else:
                    row_arr.append(convert_value(cell.value))
            data.append(row_arr)
        all_sheets[s] = data
        print(' sheet', s, 'rows', last_row, 'cols', last_col)

# Patch: don't show default Network manpower cost when no project data is entered
es = all_sheets.get('Estimate Sheet')
if es and len(es) > 79:
    # row 80, column G (index 6) is the Network Team Resource quantity
    if es[79][6] == 0.2:
        es[79][6] = "=IF(OR('Test Fit'!E4>0,'Test Fit'!E5>0),0.2,0)"

out_path = os.path.join(os.path.dirname(__file__), 'workbook.json')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(all_sheets, f, indent=None, separators=(',', ':'))
print('wrote', out_path, 'size', os.path.getsize(out_path))
